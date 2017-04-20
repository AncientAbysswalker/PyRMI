#  ____                                             _                      _           _     _              
# / ___|    ___   _ __  __   __   ___   _ __       / \     _ __     __ _  | |  _   _  | |_  (_)   ___   ___ 
# \___ \   / _ \ | '__| \ \ / /  / _ \ | '__|     / _ \   | '_ \   / _` | | | | | | | | __| | |  / __| / __|
#  ___) | |  __/ | |     \ V /  |  __/ | |       / ___ \  | | | | | (_| | | | | |_| | | |_  | | | (__  \__ \
# |____/   \___| |_|      \_/    \___| |_|      /_/   \_\ |_| |_|  \__,_| |_|  \__, |  \__| |_|  \___| |___/
#                                                                              |___/                        
#  ____   __   __       ____    __  __     ___           _                    __                       
# |  _ \  \ \ / /      |  _ \  |  \/  |   |_ _|  _ __   | |_    ___   _ __   / _|   __ _    ___    ___ 
# | |_) |  \ V /   __  | |_) | | |\/| |    | |  | '_ \  | __|  / _ \ | '__| | |_   / _` |  / __|  / _ \
# |  __/    | |   |__| |  _ <  | |  | |    | |  | | | | | |_  |  __/ | |    |  _| | (_| | | (__  |  __/
# |_|       |_|        |_| \_\ |_|  |_|   |___| |_| |_|  \__|  \___| |_|    |_|    \__,_|  \___|  \___|
#                                                                                                          
#  __  __                  _        _____       _        ___  
# |  \/  |   __ _   _ __  | | __   |___ /      / |      / _ \ 
# | |\/| |  / _` | | '__| | |/ /     |_ \      | |     | (_) |
# | |  | | | (_| | | |    |   <     ___) |  _  | |  _   \__, |
# |_|  |_|  \__,_| |_|    |_|\_\   |____/  (_) |_| (_)    /_/ 
                                                             


#http://www.messletters.com/en/big-text/
#The following are useful links that were used in initial development
#https://yuji.wordpress.com/2011/06/22/python-imaplib-imap-example-with-gmail/
#http://www.voidynullness.net/blog/2013/07/25/gmail-email-with-python-via-imap/
#http://stackoverflow.com/questions/19540192/imap-get-sender-name-and-body-text

#!/usr/bin/python2.7
# -*- coding: utf8 -*-

#Import Libraries
import imaplib, datetime
from os import path, makedirs, rename
from collections import namedtuple
#from math import ceil
from shutil import copy
now=datetime.datetime.now
today=datetime.date.today
strptime=datetime.datetime.strptime
timedelta=datetime.timedelta

def debPrint(cond,prnt): #Debug Print
	if cond:print(prnt)
	

#Define namedtuple to hold disk information
hardDisk=namedtuple('hardDisk',['Brand','Serial','Ada','errMaj','errMin','errStr','errList','timLast','timNew','timDel'])

#Missing Memory
misMem=False

##########################################################################
###   User Block

email_server='imap.gmail.com'
email_address=''
email_passkey=''
email_label='"Server/Leviathan"'  #Ensure label surrounded by both " and ' 

PyRMI_Memory="PrimaryKey.txt"
PyRMI_Summary="HumanSummary.txt"
PyRMI_RMVar="RM_LevVar.txt"
PyRMI_Trends="Server Trends.xlsx"

###
##########################################################################
###   Debugging States and Options

no_write=False
no_excel=False
print_runtime_flags=True
print_memory_read=False
print_rawdump_body=False
log_old_records=True

###
##########################################################################
###   Login and pull email data

debPrint(print_runtime_flags,"Logging into Mail Server")
mail = imaplib.IMAP4_SSL(email_server)
mail.login(email_address, email_passkey)
debPrint(print_runtime_flags,"Logged into Mail Server")

#Select folder/label to search - print(mail.list()) gives list of all "folders" aka labels if needed
mail.select(email_label) 

#result is status of search for error checking, if required for any purpose
result, data = mail.search(None, "ALL")

#Fetches raw text of latest email body with some extra crap added to it -_- Then logout of email
body=mail.fetch(data[0].split()[-1],"(UID BODY[TEXT])")[1]
mail.close()
debPrint(print_runtime_flags,"Fetched Mail. Closed connection.")
debPrint(print_rawdump_body,body)

###
##########################################################################
###   Automatic File Directories - Mess around with this at your own risk

#Define path to where this script is
PyRMIPath = path.dirname(__file__)

#Primary Runtime Files and Primary Key
MemoryPath = path.join(PyRMIPath,PyRMI_Memory)
SummaryPath = path.join(PyRMIPath,PyRMI_Summary)
RainmeterPath = path.join(PyRMIPath,PyRMI_RMVar)
ExcelPath = path.join(PyRMIPath,PyRMI_Trends)

###
##########################################################################
###   Read PyRMI Memory Document

debPrint(print_runtime_flags,"Reading PyRMI memory")
lastDate=str(today())
if path.exists(MemoryPath):
	PrimKey = open(MemoryPath,"r")
	timStmp=((PrimKey.read()).strip()).splitlines()
	[lastDate,timStmp]=[timStmp[0],timStmp[1:]]
	timStmp[:] = [disc.split("<=/=>") for disc in timStmp]
	debPrint(print_memory_read,repr(timStmp))
	PrimKey.close()
else: misMem=True

###
##########################################################################


#Locate datapool usage and convert to bytes
[_,_,SpUse,body]=((str(body).split("ZFS pool list")[1]).split("\\r\\n",3)[3]).split(None,3)	#Looks at text after "ZFS pool list" > split the body by three newlines and use the text afterwards > split the text by 3 unknown length whitespaces and return the four strings to the array
SpUse+='B'

#Locate Scrubbing Data
[_,scrubFix,_,_,_,scrubErr,body]=(body.split("scan: scrub rep")[1]).split(' ',6)	#Look for scrub flag and split with 6 whitespaces to get num of fixes and unfixable errors 

#To hold ada disk info for following code
diskInfo=[]

#Error flag in case ANY drive has a major error
errflag=0

#For each item split by SMART (some number of disks) except the first which is crap text and the last which should be da0 (USB boot)
for diskdat in body.split("S.M.A.R.T. [/dev/")[1:-1]:
	errDiscMajor=0
	errDiscMinor=0
	if diskdat[:2]=="da":
		continue
	
	discAda=diskdat[:4] #Temp var to hold ada number
	errString="" #Temp var to store all the error codes
	
	#Extracts Manufacturer and Model of HDD - Alternatively use "Device Model:" and "Serial Number:" to get model number
	discName=(diskdat.split("Model Family:")[1]).split("\\r\\n")[0][:-4].strip()
	#Extracts Serial of HDD
	discSerial=(diskdat.split("Serial Number:")[1]).split("\\r\\n")[0][:-4].strip()
	#Extracts Time Running
	#print ((((diskdat.split("9 Power_On_")[1]).split("-")[1]).split("h+")[0].strip()).split("\r\n")[0])
	#print ((((diskdat.split("9 Power_On_")[1]).split("-")[1]).split("h+")[0].strip()).splitlines(True)[0])
	discTimeNew=int((((diskdat.split("9 Power_On_")[1]).split("-")[1]).split("\\r\\n")[0]).split("h+")[0].strip())
	
	#Extract Timestamp from last running
	discTimeLast=0
	if not misMem:
		for sublist in timStmp:
			if sublist[0] == discSerial:
				discTimeLast=int(sublist[1])
	
	#Split remaining text into their respective errors and look at them all ignoring the first garbage text
	diskdat=diskdat.split("\\r\\n\\r\\nError ")
	
	errList=[]
    
	for errlog in diskdat[1:]:
	#if time of error is newer
		if int((errlog.split("lifetime: ",1)[1]).split(" hours")[0]) > discTimeLast:
			errString+="  Error " + errlog.split(' ',1)[0] + " @ " + (errlog.split("lifetime: ",1)[1]).split(')',1)[0] + ")\r\n"
			errList.append(["Error " + errlog.split(' ',1)[0]  ,  (errlog.split("lifetime: ",1)[1]).split(' hours',1)[0]])
			if errlog.split(' ',1)[0]=='5' or errlog.split(' ',1)[0]=='187' or errlog.split(' ',1)[0]=='188' or errlog.split(' ',1)[0]=='197' or errlog.split(' ',1)[0]=='198':
				errDiscMajor+=1 #Bad error found
				errflag=1
			else:
				errDiscMinor+=1
	
	#Append all data for current disk into struct
	diskInfo.append(hardDisk(discName,discSerial,discAda,errDiscMajor,errDiscMinor,errString,errList[::-1],discTimeLast,discTimeNew,discTimeNew-discTimeLast))
	debPrint(print_runtime_flags,"Resolved Disk Information:" + discAda)
diskInfo.sort() #Then sort it by Serial number - used to be by ada but ada can change
print(diskInfo)

if not no_write:
	if log_old_records and not misMem:
		debPrint(print_runtime_flags,"Logging old files")
		LogPath=path.join(PyRMIPath,"log",str(strptime(lastDate,"%Y-%m-%d").date()))
		if not path.exists(LogPath):
			makedirs(LogPath)
			rename(MemoryPath, path.join(LogPath,PyRMI_Memory))
			rename(RainmeterPath, path.join(LogPath,PyRMI_RMVar))
			rename(SummaryPath, path.join(LogPath,PyRMI_Summary))
			if not no_excel: copy(ExcelPath,path.join(LogPath,PyRMI_Trends))
		else:
			debPrint(print_runtime_flags,"Old logs already exist - did not write logs")
	
	#Load up files
	debPrint(print_runtime_flags,"Over-writing files")
	Summary = open(SummaryPath,"w")
	Rainmeter = open(RainmeterPath,"w")
	PrimKey = open(MemoryPath,"w")
	if not no_excel:
		from openpyxl import load_workbook, Workbook
		if not path.exists(ExcelPath):
			wb = Workbook();
			wb.active.title="Summary"
			wb.create_sheet("Pool")
			wb.create_sheet("Disc 1")
			wb.create_sheet("Disc 2")
			wb.create_sheet("Disc 3")
			wb.create_sheet("Disc 1 SMART")
			wb.create_sheet("Disc 2 SMART")
			wb.create_sheet("Disc 3 SMART")
						
			wb["Pool"].append(["Date", "Pool DownTime", "Scrub Fix", "Scrub Err"])
			wb["Disc 1"].append(["Date", "Uptime Error", "Min Err", "Maj Err"])
			wb["Disc 2"].append(["Date", "Uptime Error", "Min Err", "Maj Err"])
			wb["Disc 3"].append(["Date", "Uptime Error", "Min Err", "Maj Err"])
			wb["Disc 1 SMART"].append(["CODE","Time","Hours"])
			wb["Disc 2 SMART"].append(["CODE","Time","Hours"])
			wb["Disc 3 SMART"].append(["CODE","Time","Hours"])
		else:
			wb = load_workbook(ExcelPath)

	#Write all data to the rainmeter data file or the summary file
	Rainmeter.write("[Variables]\r\npoolUse=" + str(SpUse) + "\r\nscrubFix=" + scrubFix + "\r\nscrubErr=" + scrubErr + '\r\nlastDate="' + today().strftime("%B %d, %Y") + '"')
	Summary.write("Scrubing of datapool repaired " + scrubFix + " and found " + scrubErr + " to be unrepairable.\r\n")


	maxTime=max(disks.timDel for disks in diskInfo)

	if not no_excel:
		debPrint(print_runtime_flags,"Appending data to excel")
		if not misMem:
			wkDiff=today().isocalendar()[1]-strptime(lastDate,"%Y-%m-%d").date().isocalendar()[1]+12*(today().isocalendar()[0]-strptime(lastDate,"%Y-%m-%d").date().isocalendar()[0]) 
			wb["Pool"].append([strptime(lastDate,"%Y-%m-%d").date(), 168*wkDiff-maxTime, int(scrubFix), int(scrubErr)])
			wb["Pool"].append([today(), 168*wkDiff-maxTime, int(scrubFix), int(scrubErr)])
		wb["Pool"].append([today(), 0, 0, 0])

	#Write SMART output for RM
	if errflag==1:
		Rainmeter.write('\r\nSMART="Serious problems with pool disks were identified"')
	else:
		Rainmeter.write('\r\nSMART="No serious problems were found with pool disks"')
	
	iter=1
	PrimKey.write(str(today()) + "\r\n")
	for disks in diskInfo:
		Summary.write("<" + disks.Brand.split(' ')[0] + " " + disks.Serial + ">-------------------------------------------<" + disks.Ada + ">\r\n" + disks.errStr + "\r\n")
		PrimKey.write(disks.Serial + "<=/=>" + str(disks.timNew) + "\r\n")
	
		if not no_excel:
			if not misMem:
				wb["Disc " + str(iter)].append([strptime(lastDate,"%Y-%m-%d").date(), maxTime-disks.timDel, disks.errMin, disks.errMaj])
				wb["Disc " + str(iter)].append([today(), maxTime-disks.timDel, disks.errMin, disks.errMaj])
			wb["Disc " + str(iter)].append([today(), 0, 0, 0])
			for errors in disks.errList:
				wb["Disc " + str(iter) + " SMART"].append(['','',errors[0],(now()-timedelta(days=7)+timedelta(hours=(int(errors[1])-disks.timLast))).strftime("%b %d, %Y %H") + u":00 +/- " + str(0) + "hours",errors[1]])
		iter+=1

Summary.close()
Rainmeter.close()
PrimKey.close()
if not no_excel: wb.save(ExcelPath)
debPrint(print_runtime_flags,"Closing - Done")
